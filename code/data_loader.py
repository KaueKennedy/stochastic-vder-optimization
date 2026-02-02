import pandas as pd
import numpy as np
import sys
import os
import random
import networkx as nx
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
from scipy.spatial.distance import cdist
from config import CFG
import os
import os

# ==============================================================================
# CONFIGURATION (CAMINHOS DINÂMICOS E BLINDADOS)
# ==============================================================================
# 1. Descobre onde o arquivo 'visualizer.py' está salvo no disco
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 2. Sobe um nível para chegar na pasta raiz do projeto (pasta 'codes')
# Ele faz: "C:\...\code" -> "C:\...\codes"
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

# 3. Agora montamos os caminhos a partir da raiz segura
# (Não importa onde você dê o Play, ele sempre vai acertar)
DATA_FILE = os.path.join(PROJECT_ROOT, "resultados", "RELATORIO_COMPLETO_SCORES.csv")
BUS_FILE = os.path.join(PROJECT_ROOT, "resultados", "MASTER_Bus_Results.csv")

# Excel da Rede (Note como é fácil navegar nas pastas agora)
EXCEL_REDE = os.path.join(PROJECT_ROOT, "Iowa_Distribution_Test_Systems", "OpenDSS Model", "OpenDSS Model", "Rede_240Bus_Dados.xlsx")

# Arquivos de Contexto
CONTEXT_FILES = {
    "Hourly": os.path.join(PROJECT_ROOT, "resultados", "MASTER_Hourly_Results.csv"),
    "Bus": BUS_FILE,
    "Voltage": os.path.join(PROJECT_ROOT, "resultados", "MASTER_Voltage_Log.csv"),
    "History": os.path.join(PROJECT_ROOT, "resultados", "Optimization_History_Log.csv")
}

# ==============================================================================
# CONFIGURATION (CAMINHOS DINÂMICOS E BLINDADOS)
# ==============================================================================
# 1. Descobre onde o arquivo 'visualizer.py' está salvo no disco
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 2. Sobe um nível para chegar na pasta raiz do projeto (pasta 'codes')
# Ele faz: "C:\...\code" -> "C:\...\codes"
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

# 3. Agora montamos os caminhos a partir da raiz segura
# (Não importa onde você dê o Play, ele sempre vai acertar)
DATA_FILE = os.path.join(PROJECT_ROOT, "resultados", "RELATORIO_COMPLETO_SCORES.csv")
BUS_FILE = os.path.join(PROJECT_ROOT, "resultados", "MASTER_Bus_Results.csv")

# Excel da Rede (Note como é fácil navegar nas pastas agora)
EXCEL_REDE = os.path.join(PROJECT_ROOT, "Iowa_Distribution_Test_Systems", "OpenDSS Model", "OpenDSS Model", "Rede_240Bus_Dados.xlsx")

# Arquivos de Contexto
CONTEXT_FILES = {
    "Hourly": os.path.join(PROJECT_ROOT, "resultados", "MASTER_Hourly_Results.csv"),
    "Bus": BUS_FILE,
    "Voltage": os.path.join(PROJECT_ROOT, "resultados", "MASTER_Voltage_Log.csv"),
    "History": os.path.join(PROJECT_ROOT, "resultados", "Optimization_History_Log.csv")
}

# ==============================================================================
# FUNÇÃO AUXILIAR: CLASSIFICAÇÃO TOPOLÓGICA (VIA CARGA E DISTÂNCIA)
# ==============================================================================
def classify_zones_topological(buses_df, loads_df, lines_df):
    """
    Classifica zonas baseando-se EXCLUSIVAMENTE em Agrupamentos de Carga.
    A Subestação NÃO é considerada centro urbano, a menos que haja carga nela.
    """
    print("   -> [Zone Classifier] Detecting Load Centers (Ignoring Substation location)...")
    
    # --- 1. PREPARAÇÃO DOS DADOS ---
    def clean_name(name):
        n = str(name).lower().strip()
        for tag in ['t_', '_l', '_node', 'load_']: n = n.replace(tag, '')
        return n

    # Remove qualquer zona pré-existente para não confundir
    if 'Zone' in buses_df.columns: buses_df = buses_df.drop(columns=['Zone'])
    
    # Cria coluna limpa para join
    buses_df['Bus_Clean'] = buses_df['BusName'].apply(clean_name)
    
    # Filtro de Coordenadas: Remove quem não tem GPS
    mask_valid = (buses_df['x'].abs() > 0.001) | (buses_df['y'].abs() > 0.001)
    df_geo = buses_df[mask_valid].copy()
    valid_names = set(df_geo['Bus_Clean'].unique())

    # Mapear Cargas
    node_loads = {n: 0.0 for n in valid_names}
    if loads_df is not None:
        for _, row in loads_df.iterrows():
            bus_raw = clean_name(row['Bus'])
            kw = float(row['kW'])
            if bus_raw in node_loads: node_loads[bus_raw] += kw

    df_geo['Load_kW'] = df_geo['Bus_Clean'].map(node_loads)

    # --- 2. K-MEANS ESPACIAL (AGRUPAR BAIRROS) ---
    N_CLUSTERS = 10
    coords = df_geo[['x', 'y']].values
    
    if len(coords) < N_CLUSTERS: N_CLUSTERS = len(coords)
    
    kmeans = KMeans(n_clusters=N_CLUSTERS, random_state=42, n_init=10)
    df_geo['Cluster_ID'] = kmeans.fit_predict(coords)
    
    # --- 3. IDENTIFICAR CLUSTERS URBANOS (SOMENTE POR CARGA) ---
    cluster_stats = df_geo.groupby('Cluster_ID')['Load_kW'].sum().reset_index()
    cluster_stats.columns = ['Cluster_ID', 'Total_Load']
    
    mu = cluster_stats['Total_Load'].mean()
    sigma = cluster_stats['Total_Load'].std()
    
    # Regra: Só é centro urbano se a carga do cluster for alta
    threshold = mu + (0.5 * sigma)
    urban_clusters = cluster_stats[cluster_stats['Total_Load'] > threshold]['Cluster_ID'].tolist()
    
    print(f"      [Zoning] {len(urban_clusters)} Load Centers identified (Load > {threshold:.1f} kW)")
    
    # --- 4. DEFINIR ÂNCORAS (CENTROS DAS CIDADES) ---
    anchor_nodes = []
    anchor_region_loads = {} 
    centers = kmeans.cluster_centers_
    
    for cid in urban_clusters:
        centroid = centers[cid]
        cluster_total_kw = cluster_stats.loc[cluster_stats['Cluster_ID'] == cid, 'Total_Load'].values[0]
        
        # Acha o bus físico mais próximo do centro matemático do cluster
        cluster_buses = df_geo[df_geo['Cluster_ID'] == cid]
        if not cluster_buses.empty:
            c_coords = cluster_buses[['x', 'y']].values
            dists = cdist([centroid], c_coords)
            nearest_idx = dists.argmin()
            nearest_bus = cluster_buses.iloc[nearest_idx]['Bus_Clean']
            
            if nearest_bus not in anchor_nodes:
                anchor_nodes.append(nearest_bus)
                anchor_region_loads[nearest_bus] = cluster_total_kw

    # --- 5. CÁLCULO DE SCORE DE DISTÂNCIA ---
    G = nx.Graph()
    G.add_nodes_from(valid_names)
    for _, row in lines_df.iterrows():
        u, v = clean_name(row['Bus1']), clean_name(row['Bus2'])
        if u in valid_names and v in valid_names: G.add_edge(u, v, weight=1.0)
    
    try:
        bet_map = nx.betweenness_centrality(G, weight='weight')
    except:
        bet_map = {n:0 for n in G.nodes}

    node_scores = []
    for n in G.nodes:
        min_dist = 9999
        if not anchor_nodes:
            min_dist = 100 
        else:
            for anchor in anchor_nodes:
                try:
                    d = nx.shortest_path_length(G, source=n, target=anchor)
                    if d < min_dist: min_dist = d
                except: pass
        
        if min_dist == 9999: min_dist = 50
        
        node_scores.append({
            'Bus': n,
            'Dist_Min': min_dist,
            'Bet': bet_map.get(n, 0)
        })

    df_topo = pd.DataFrame(node_scores)
    
    if not df_topo.empty:
        scaler = MinMaxScaler()
        df_topo[['Dist_Norm', 'Bet_Norm']] = scaler.fit_transform(df_topo[['Dist_Min', 'Bet']])
        
        # Score: 60% Proximidade Urbana, 40% Centralidade
        df_topo['Raw_Score'] = ((1 - df_topo['Dist_Norm']) * 0.6) + (df_topo['Bet_Norm'] * 0.4)
        
        scaler100 = MinMaxScaler(feature_range=(0, 100))
        df_topo['Final_Score'] = scaler100.fit_transform(df_topo[['Raw_Score']])

        # --- 6. CLASSIFICAÇÃO ---
        p_urban = np.percentile(df_topo['Final_Score'], 85)
        p_mixed = np.percentile(df_topo['Final_Score'], 50)
        
        def get_zone(score):
            if score >= p_urban: return 'Urban'
            elif score >= p_mixed: return 'Mixed'
            else: return 'Rural'

        df_topo['Zone'] = df_topo['Final_Score'].apply(get_zone)
        # Garante âncoras como Urban
        df_topo.loc[df_topo['Bus'].isin(anchor_nodes), 'Zone'] = 'Urban'
    else:
        # Fallback se não houver topologia válida
        df_topo = pd.DataFrame({'Bus': list(valid_names), 'Zone': 'Mixed', 'Final_Score': 50})

    # Merge para trazer a zona calculada para perto das coordenadas
    df_merged = buses_df.merge(
        df_topo[['Bus', 'Zone', 'Final_Score']], 
        left_on='Bus_Clean', right_on='Bus', how='left'
    )
    
    # Exemplo de correção rápida antes do plot
    # Se t_bus1003_l não tem X, pega o X de bus1003
    for idx, row in buses_df.iterrows():
        if (abs(row['x']) < 0.001) and ('t_' in row['BusName']):
            # Tenta achar o 'pai' limpando o nome
            parent_name = row['BusName'].replace('t_', '').replace('_l', '')
            parent_data = buses_df[buses_df['BusName'] == parent_name]
            
            if not parent_data.empty:
                buses_df.at[idx, 'x'] = parent_data.iloc[0]['x']
                buses_df.at[idx, 'y'] = parent_data.iloc[0]['y']

    # Quem ficou de fora (sem coordenada ou desconexo) vira Rural
    df_merged['Zone'] = df_merged['Zone'].fillna('Rural')

    # ---------------------------------------------------------
    # 7. PLOTAGEM (Usa df_merged, que sabemos estar CORRETO)
    # ---------------------------------------------------------
    coords_map = {}
    for _, r in df_geo.iterrows():
        coords_map[r['Bus_Clean']] = (r['x'], r['y'])

    # Edges
    edge_x, edge_y = [], []
    for u, v in G.edges():
        if u in coords_map and v in coords_map:
            x0, y0 = coords_map[u]
            x1, y1 = coords_map[v]
            edge_x.extend([x0, x1, None])
            edge_y.extend([y0, y1, None])

    edge_trace = go.Scatter(
        x=edge_x, y=edge_y,
        line=dict(width=1, color='#bdc3c7'), hoverinfo='none', mode='lines'
    )

    fig = go.Figure(data=[edge_trace])
    colors = {'Urban': '#e74c3c', 'Mixed': '#f1c40f', 'Rural': '#2ecc71'}

    # Plot Nodes
    for zone in ['Rural', 'Mixed', 'Urban']:
        subset = df_merged[(df_merged['Zone'] == zone) & (~df_merged['Bus_Clean'].isin(anchor_nodes))]
        if not subset.empty:
            nx_vals, ny_vals, texts = [], [], []
            for _, row in subset.iterrows():
                bname = row['Bus_Clean']
                if bname in coords_map:
                    x, y = coords_map[bname]
                    nx_vals.append(x); ny_vals.append(y)
                    score = row['Final_Score'] if pd.notnull(row['Final_Score']) else 0
                    ind_load = node_loads.get(bname, 0.0)
                    texts.append(f"Bus: {row['BusName']}<br>Zone: {zone}<br>Score: {score:.1f}<br>Load: {ind_load:.2f} kW")
            
            if nx_vals:
                fig.add_trace(go.Scatter(
                    x=nx_vals, y=ny_vals, mode='markers', name=f"{zone} Zone", text=texts, hoverinfo='text',
                    marker=dict(color=colors[zone], size=8 if zone=='Rural' else 10)
                ))

    # Plot ÂNCORAS
    ax, ay, atxt = [], [], []
    for anchor in anchor_nodes:
        if anchor in coords_map:
            x, y = coords_map[anchor]
            ax.append(x); ay.append(y)
            reg_load = anchor_region_loads.get(anchor, 0.0)
            atxt.append(f"<b>URBAN CENTER</b><br>{anchor}<br>Region Load: {reg_load:,.2f} kW")
            
    if ax:
        fig.add_trace(go.Scatter(
            x=ax, y=ay, mode='markers', name='Urban Centers', text=atxt, hoverinfo='text',
            marker=dict(symbol='star', color='black', size=20, line=dict(width=2, color='white'))
        ))

    # Círculos Visuais
    for cid in urban_clusters:
        cx, cy = centers[cid]
        fig.add_trace(go.Scatter(
            x=[cx], y=[cy], mode='markers', showlegend=False,
            marker=dict(size=80, color='rgba(231, 76, 60, 0.15)', line=dict(width=0)),
            hoverinfo='none'
        ))

    fig.update_layout(
        title=f"Zoning (Based on Load Centers)",
        showlegend=True, hovermode='closest',
        margin=dict(b=20,l=20,r=20,t=40),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, scaleanchor="x", scaleratio=1),
        plot_bgcolor='white'
    )
    
    # ==============================================================================
    # 8. CONSOLIDAÇÃO DA VERDADE (ATUALIZANDO O DATAFRAME)
    # ==============================================================================
    print("      [Zone Classifier] Consolidating Data used in Graph...")
    
    # Cria mapa: Nome_Limpo -> Zona (Ex: 'bus1003' -> 'Rural')
    truth_map = dict(zip(df_merged['Bus_Clean'], df_merged['Zone']))
    
    # Aplica no DataFrame original. 
    # A MÁGICA ACONTECE AQUI:
    # Como aplicamos sobre 'Bus_Clean', tanto 'bus1003' quanto 't_bus1003_l' 
    # terão o mesmo 'Bus_Clean' e receberão a mesma zona.
    buses_df['Zone'] = buses_df['Bus_Clean'].map(truth_map).fillna('Rural')
    
    final_counts = buses_df['Zone'].value_counts().to_dict()
    print(f"      [Zone Classifier] ✅ ZONAS CONFIRMADAS: {final_counts}")

    # ==============================================================================
    # 9. SALVAR NO EXCEL (A PARTE QUE VOCÊ PEDIU)
    # ==============================================================================
    # Caminho fixo conforme solicitado
    excel_path = r".\Iowa_Distribution_Test_Systems\OpenDSS Model\OpenDSS Model\Rede_240Bus_Dados.xlsx"
    
    # Importação local para garantir que funcione mesmo se esquecer lá em cima
    try:
        from openpyxl import load_workbook
        print(f"💾 Salvando atualizações no arquivo: {excel_path}")
        
        # 1. Carrega o Excel mantendo as fórmulas e formatações
        wb = load_workbook(excel_path)
        
        if 'Buses' in wb.sheetnames:
            ws = wb['Buses']
            
            # 2. Cria dicionário de busca rápida: {Nome_Exato_No_Excel : Zona_Calculada}
            # O buses_df já tem a zona correta para 't_bus1003_l' graças ao passo 8.
            zone_lookup = dict(zip(buses_df['BusName'], buses_df['Zone']))
            
            count_update = 0
            
            # 3. Itera sobre as linhas do Excel (Começando da linha 2)
            # Coluna A (idx 0) = BusName
            # Coluna E (idx 4) = Zone
            for row in ws.iter_rows(min_row=2):
                cell_name = row[0] # Coluna A
                cell_zone = row[4] # Coluna E
                
                if cell_name.value:
                    xls_name = str(cell_name.value).strip()
                    
                    # Se esse nome existe nos nossos cálculos, atualiza a zona
                    if xls_name in zone_lookup:
                        cell_zone.value = zone_lookup[xls_name]
                        count_update += 1
            
            # 4. Salva o arquivo
            wb.save(excel_path)
            print(f"✅ EXCEL ATUALIZADO: {count_update} linhas modificadas na aba 'Buses'.")
        else:
            print(f"❌ Erro: Aba 'Buses' não encontrada no Excel.")
            
    except Exception as e:
        print(f"❌ FALHA AO SALVAR EXCEL: {e}")
        print("⚠️  AVISO: Feche o arquivo Excel se ele estiver aberto!")

    return buses_df, fig

# ==============================================================================
# CARREGAMENTO PRINCIPAL
# ==============================================================================
def load_inputs():
    print("--- [1] Carregando Dados (Módulo DataLoader) ---")
    
    if not os.path.exists(CFG["EXCEL_REDE"]):
        print(f"ERRO CRÍTICO: Excel não encontrado em {CFG['EXCEL_REDE']}")
        sys.exit(1)

    try:
        df_lines = pd.read_excel(CFG["EXCEL_REDE"], sheet_name="Lines")
        df_buses = pd.read_excel(CFG["EXCEL_REDE"], sheet_name="Buses")
        df_loads = pd.read_excel(CFG["EXCEL_REDE"], sheet_name="Loads")
        
        # Normalização Básica
        if 'BusName' in df_buses.columns:
            df_buses['BusName'] = df_buses['BusName'].astype(str).str.strip()
        df_lines['Bus1'] = df_lines['Bus1'].astype(str).str.strip()
        df_lines['Bus2'] = df_lines['Bus2'].astype(str).str.strip()
        df_loads['Bus'] = df_loads['Bus'].astype(str).str.strip()

        # CLASSIFICAÇÃO TOPOLÓGICA
        df_buses, fig_zones = classify_zones_topological(df_buses, df_loads, df_lines)
        
        #try:
        #    st.sidebar.markdown("### 🗺️ Network Zones")
        #    st.sidebar.plotly_chart(fig_zones, width="stretch")
        #except: pass 

        print("   -> Gerando Scores de Vulnerabilidade Social...")
        scores = []
        social_cfg = CFG.get("SOCIAL_SCORING", {})
        default_cfg = social_cfg.get("Default", {"base": 2.0, "sigma": 1.0})

        for z in df_buses['Zone']:
            params = social_cfg.get(z, default_cfg)
            base = params["base"]
            sigma = params["sigma"]
            s = random.gauss(base, sigma)
            s = max(0, min(10, s))
            scores.append(s)
        
        df_buses['SocialScore'] = scores
    
        print(f"   -> Excel lido com sucesso ({len(df_loads)} cargas).")
    except Exception as e:
        print(f"ERRO ao ler abas do Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    def get_prof(path, name):
        if not os.path.exists(path):
            print(f"ERRO: Arquivo de perfil não encontrado: {path}")
            sys.exit(1)
        try:
            df = pd.read_csv(path, header=None)
            if isinstance(df.iloc[0,0], str): 
                try: float(df.iloc[0,0])
                except: df = pd.read_csv(path)
            vals = df.select_dtypes(include=[np.number]).iloc[:,0].values
            return np.resize(vals, CFG['T'])[:CFG['T']]
        except Exception as e:
            print(f"ERRO CRÍTICO ao ler perfil '{name}': {e}")
            sys.exit(1)

    load_curve = get_prof(CFG["FILE_LOAD"], "Carga")
    solar = get_prof(CFG["FILE_SOLAR"], "Solar")
    wind = get_prof(CFG["FILE_WIND"], "Eólica")
    price = get_prof(CFG["FILE_PRICE"], "Preço")
    price = price / 1000.0

    if np.max(load_curve) > 10: load_curve /= np.max(load_curve)
    if np.max(solar) > 10: solar /= np.max(solar)

    try:
        threshold_pct = CFG["REMUNERATION"].get("PEAK_THRESHOLD_PCT", 0.85)
        limit_val = np.quantile(load_curve, threshold_pct)
        if limit_val == np.min(load_curve): limit_val = np.max(load_curve) * 0.95
        dynamic_peaks = np.where(load_curve >= limit_val)[0].tolist()
        CFG["REMUNERATION"]["PEAK_HOURS"] = dynamic_peaks
        print(f"   -> [Daily Peak] Critério: Carga > {limit_val:.3f} kW. Horas: {dynamic_peaks}")
    except Exception as e:
        print(f"   [AVISO] Falha no cálculo de pico: {e}")
        dynamic_peaks = []

    profiles_dict = {
        'solar': solar,
        'wind': wind,
        'price': price,
        'load_curve': load_curve,
        'peak_hours_indices': dynamic_peaks
    }

    return df_lines, df_buses, df_loads, profiles_dict, fig_zones

# Em data_loader.py
def generate_stochastic_scenarios(base_profiles, stochastic_cfg):
    """
    Gera cenários baseados na configuração dinâmica do Dashboard.
    """
    S = stochastic_cfg["NUM_SCENARIOS"]
    T = len(base_profiles['load_curve'])
    sigma_load = stochastic_cfg["SIGMA_LOAD"]
    sigma_ren = stochastic_cfg["SIGMA_RENEWABLE"]
    
    rng = np.random.default_rng(42) # Seed fixa para reprodutibilidade visual
    
    scenarios = []
    for s in range(S):
        # Ruído multiplicativo centrado em 1.0
        l_noise = rng.normal(1.0, sigma_load, T)
        s_noise = rng.normal(1.0, sigma_ren, T)
        
        scenarios.append({
            'id': s,
            'prob': 1.0 / S,
            # Garante não negativo com np.maximum(0, ...)
            'load': np.maximum(0, base_profiles['load_curve'] * l_noise),
            'solar': np.maximum(0, base_profiles['solar'] * s_noise),
            'wind': np.maximum(0, base_profiles['wind'] * s_noise),
            'price': base_profiles['price'] # Preço geralmente mantemos fixo ou criamos outro sigma
        })
    
    return scenarios